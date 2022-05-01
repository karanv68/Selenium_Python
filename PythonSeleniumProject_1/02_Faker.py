# Faker generates the fake data for users

from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
fake_data = Faker(['hi_IN'])  # in [''] it is a locale to give the data in particular language

#print(fake_data.name()) # it will create the fake name 
#print(fake_data.email()) # it will print the fake email
#print(fake_data.address()) # it will print the fake address

for i in range(1,11):
    for j in range(1,4):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email()
        ws.cell(row=i, column=3).value = fake_data.address()

wb.save("Testdata.xlsx")


